import os
import re
import time
from typing import Dict, List, Optional, Callable, Any

import pandas as pd
from dotenv import load_dotenv
from serpapi import GoogleSearch
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from io import BytesIO


# ----------------------------
# ENV / KEY
# ----------------------------
load_dotenv()
_DEFAULT_KEY = os.getenv("SERPAPI_KEY")


def load_key_or_raise(override_key: Optional[str] = None) -> str:
    key = override_key or _DEFAULT_KEY
    if not key or not key.strip():
        raise ValueError("Missing SERPAPI_KEY. Put it in .env (SERPAPI_KEY=...) or input in UI.")
    return key.strip()


# ----------------------------
# City coordinates (ll)
# ----------------------------
# Bạn có thể tinh chỉnh zoom (13z/12z/11z) nếu muốn thu hẹp/mở rộng.
CITY_LL = {
    "TP.HCM": "@10.776889,106.700806,13z",
    "Hà Nội": "@21.027763,105.834160,13z",
    "Hải Phòng": "@20.844911,106.688084,12z",
    "Quảng Ninh": "@21.006382,107.292514,11z",  # gần Hạ Long
}

# Khu vực (để xuất sheet theo miền)
REGION_BY_CITY = {
    "Hà Nội": "Miền Bắc",
    "Hải Phòng": "Miền Bắc",
    "Quảng Ninh": "Miền Bắc",
    "TP.HCM": "Miền Nam",
}

# District packs (giúp lên số lượng nhanh hơn)
# Lưu ý: Đây là danh sách cấp huyện/quận phổ biến để tạo query theo địa bàn.
DISTRICTS = {
    "TP.HCM": [
        "Quận 1", "Quận 3", "Quận 5", "Quận 7", "Quận 10",
        "Bình Thạnh", "Phú Nhuận", "Thủ Đức", "Tân Bình", "Gò Vấp"
    ],
    "Hà Nội": [
        "Hoàn Kiếm", "Ba Đình", "Đống Đa", "Cầu Giấy", "Thanh Xuân",
        "Hai Bà Trưng", "Tây Hồ", "Nam Từ Liêm", "Bắc Từ Liêm", "Long Biên"
    ],
    "Hải Phòng": [
        "Hồng Bàng", "Ngô Quyền", "Lê Chân", "Kiến An",
        "Hải An", "Đồ Sơn", "Dương Kinh", "An Dương",
        "Thủy Nguyên", "An Lão", "Kiến Thụy", "Tiên Lãng",
        "Vĩnh Bảo", "Cát Hải", "Bạch Long Vĩ"
    ],
    "Quảng Ninh": [
        "Hạ Long", "Cẩm Phả", "Uông Bí", "Móng Cái",
        "Đông Triều", "Quảng Yên",
        "Hải Hà", "Tiên Yên", "Vân Đồn", "Đầm Hà",
        "Bình Liêu", "Ba Chẽ", "Cô Tô"
    ],
}

BRANDS_CAFE = [
    "The Coffee House", "Highlands Coffee", "Phúc Long", "Phuc Long",
    "Starbucks", "Trung Nguyên", "Katinat", "Cheese Coffee", "Cộng Cà Phê"
]
BRANDS_COWORK = [
    "Toong", "Dreamplex", "CirCO", "Circo", "Regus", "UPGen", "The Hive", "WeWork"
]

# Core queries
CORE_QUERIES_COWORK = [
    "coworking space {city}",
    "shared office {city}",
    "serviced office {city}",
    "flexible workspace {city}",
    "không gian làm việc chung {city}",
    "văn phòng chia sẻ {city}",
]
CORE_QUERIES_CAFE = [
    "quán cà phê làm việc {city}",
    "cafe làm việc {city}",
    "cafe yên tĩnh {city}",
    "cafe có ổ cắm {city}",
    "cafe học bài {city}",
    "cafe wifi mạnh {city}",
    "coffee shop good for working {city}",
    "quiet cafe {city}",
]


# ----------------------------
# Helpers
# ----------------------------
def _norm(s: str) -> str:
    s = (s or "").strip().lower()
    s = re.sub(r"\s+", " ", s)
    return s


def detect_brand(title: str) -> Optional[str]:
    t = (title or "").lower()
    for b in BRANDS_CAFE + BRANDS_COWORK:
        if b.lower() in t:
            return b
    return None


def classify_place(types: List[str], type_text: str, title: str) -> str:
    hay = " ".join([*(types or []), type_text or "", title or ""]).lower()

    # coworking signals
    if any(k in hay for k in [
        "cowork", "shared office", "serviced office", "workspace",
        "văn phòng", "không gian làm việc chung"
    ]):
        return "coworking space"

    # cafe signals
    if any(k in hay for k in ["cafe", "coffee", "quán cà phê", "coffee_shop", "coffee shop"]):
        return "quán cà phê làm việc"

    # fallback by brand
    b = detect_brand(title)
    if b and b in BRANDS_COWORK:
        return "coworking space"
    if b and b in BRANDS_CAFE:
        return "quán cà phê làm việc"

    return "khác"


def extract_notes(extensions) -> str:
    """
    Gợi ý ghi chú: lấy các tín hiệu liên quan làm việc từ extensions
    """
    notes = []
    if isinstance(extensions, list):
        for block in extensions:
            if not isinstance(block, dict):
                continue
            for k, v in block.items():
                if isinstance(v, list):
                    vv = " | ".join(str(x) for x in v)
                    if any(sig in vv.lower() for sig in [
                        "working", "laptop", "wi-fi", "wifi", "quiet", "yên tĩnh",
                        "cozy", "outlet", "ổ cắm"
                    ]):
                        notes.append(f"{k}: {vv}")
    out = " ; ".join(notes)
    return out[:500]


def maps_link_from_place_id(place_id: Optional[str]) -> Optional[str]:
    if not place_id:
        return None
    return f"https://www.google.com/maps/place/?q=place_id:{place_id}"

def extract_image_url(r: dict) -> str:
    """
    SerpApi local_results thường có thumbnail hoặc photo.
    Trả về URL ảnh nếu có, không có thì None.
    """
    for k in ["thumbnail", "image", "photo", "thumbnail_url"]:
        v = r.get(k)
        if isinstance(v, str) and v.startswith("http"):
            return v

    # đôi khi photos là list dict
    photos = r.get("photos")
    if isinstance(photos, list) and photos:
        first = photos[0]
        if isinstance(first, dict):
            for kk in ["thumbnail", "image", "url"]:
                vv = first.get(kk)
                if isinstance(vv, str) and vv.startswith("http"):
                    return vv

    return None


def _safe_sheet_name(name: str) -> str:
    """
    Excel sheet name constraints:
    - max 31 chars
    - cannot contain: : \ / ? * [ ]
    """
    if not name:
        return "Sheet"
    n = re.sub(r'[:\\/?*\[\]]+', " ", name).strip()
    n = re.sub(r"\s+", " ", n)
    return n[:31] if len(n) > 31 else n


# ----------------------------
# Query Builder
# ----------------------------
def build_queries(
    cities: List[str],
    use_core: bool = True,
    use_brands: bool = True,
    use_districts: bool = True,
) -> Dict[str, List[str]]:
    queries_by_city: Dict[str, List[str]] = {}

    for city in cities:
        qs: List[str] = []

        if use_core:
            qs += [q.format(city=city) for q in CORE_QUERIES_COWORK]
            qs += [q.format(city=city) for q in CORE_QUERIES_CAFE]

        if use_brands:
            qs += [f"{b} {city}" for b in BRANDS_COWORK]
            qs += [f"{b} {city}" for b in BRANDS_CAFE]

        if use_districts:
            dists = DISTRICTS.get(city, [])
            for d in dists:
                qs += [
                    f"coworking space {d} {city}",
                    f"không gian làm việc chung {d} {city}",
                    f"quán cà phê làm việc {d} {city}",
                    f"cafe làm việc {d} {city}",
                    f"cafe có ổ cắm {d} {city}",
                    f"cafe yên tĩnh {d} {city}",
                    f"cafe học bài {d} {city}",
                ]

        # unique while keep order
        seen = set()
        uniq = []
        for q in qs:
            k = _norm(q)
            if k not in seen:
                seen.add(k)
                uniq.append(q)
        queries_by_city[city] = uniq

    return queries_by_city


# ----------------------------
# SerpApi call
# ----------------------------
def serpapi_maps_search(
    api_key: str,
    q: str,
    ll: str,
    start: int,
    hl: str = "vi",
    gl: str = "vn",
    google_domain: str = "google.com",
) -> Dict[str, Any]:
    params = {
        "engine": "google_maps",
        "type": "search",
        "q": q,
        "ll": ll,
        "hl": hl,
        "gl": gl,
        "google_domain": google_domain,
        "start": start,
        "api_key": api_key,
    }
    return GoogleSearch(params).get_dict()


# ----------------------------
# Crawler
# ----------------------------
def crawl_cities(
    cities: List[str],
    queries_by_city: Dict[str, List[str]],
    max_pages_per_query: int = 10,
    sleep_sec: float = 0.3,
    min_rating: float = 0.0,
    min_reviews: int = 0,
    on_progress: Optional[Callable[[str, int, int], None]] = None,
    should_stop: Optional[Callable[[], bool]] = None,
    override_key: Optional[str] = None,
) -> List[Dict[str, Any]]:
    api_key = load_key_or_raise(override_key)

    all_rows: List[Dict[str, Any]] = []
    global_seen = set()

    total_queries = sum(len(queries_by_city.get(c, [])) for c in cities)
    done_queries = 0

    for city in cities:
        ll = CITY_LL.get(city)
        if not ll:
            raise ValueError(f"Missing ll for city: {city}")

        queries = queries_by_city.get(city, [])
        for q in queries:
            if should_stop and should_stop():
                return all_rows

            done_queries += 1
            if on_progress:
                on_progress(f"Đang crawl: {city} | Query: {q}", done_queries, total_queries)

            for page in range(max_pages_per_query):
                if should_stop and should_stop():
                    return all_rows

                start = page * 20
                data = serpapi_maps_search(api_key=api_key, q=q, ll=ll, start=start)
                local_results = data.get("local_results", []) or []

                if not local_results:
                    break

                new_in_page = 0
                for r in local_results:
                    data_id = r.get("data_id")
                    place_id = r.get("place_id")
                    cid = r.get("data_cid")

                    title = r.get("title")
                    address = r.get("address") or r.get("Địa chỉ")

                    # key chống trùng
                    key = data_id or place_id or cid or (_norm(title) + "|" + _norm(address))
                    if key in global_seen:
                        continue

                    gps = r.get("gps_coordinates") or {}
                    lat = gps.get("latitude")
                    lng = gps.get("longitude")

                    rating = r.get("rating")
                    reviews_count = r.get("reviews")

                    # lọc rating/reviews (nếu người dùng bật)
                    if min_rating and isinstance(rating, (int, float)) and rating < min_rating:
                        continue
                    if min_reviews and isinstance(reviews_count, int) and reviews_count < min_reviews:
                        continue

                    types = r.get("types") or []
                    type_text = r.get("type") or ""
                    brand = detect_brand(title)
                    category = classify_place(types=types, type_text=type_text, title=title)

                    notes = extract_notes(r.get("extensions"))
                    website = r.get("website")
                    image_url = extract_image_url(r)
                    row = {
                        # mapping chuẩn
                        "name": title,
                        "category": category,
                        "brand": brand,
                        "address": address,
                        "city": city,
                        "region": REGION_BY_CITY.get(city, ""),
                        "google_maps_link": maps_link_from_place_id(place_id),
                        "website_or_social": website,
                        "latitude": lat,
                        "longitude": lng,
                        "notes": notes,
                        "image_url": image_url,
                        # raw thêm để audit
                        "phone": r.get("phone"),
                        "rating": rating,
                        "reviews_count": reviews_count,
                        "data_id": data_id,
                        "place_id": place_id,
                        "data_cid": cid,
                        "source_query": q,
                    }

                    global_seen.add(key)
                    all_rows.append(row)
                    new_in_page += 1

                # nếu page này gần như không có mới => dừng sớm tiết kiệm quota
                if new_in_page <= 2 and page >= 2:
                    break

                time.sleep(sleep_sec)

    return all_rows


# ----------------------------
# Postprocess
# ----------------------------
def postprocess_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return pd.DataFrame(columns=[
            "name", "category", "brand", "address", "city", "region",
            "google_maps_link", "website_or_social",
            "latitude", "longitude", "notes", "image_url",
        ])

    df = df.copy()

    # normalize empties
    df["name"] = df.get("name", "").fillna("").astype(str).str.strip()
    df["address"] = df.get("address", "").fillna("").astype(str).str.strip()
    df["city"] = df.get("city", "").fillna("").astype(str).str.strip()

    # ensure region exists
    if "region" not in df.columns:
        df["region"] = df["city"].map(REGION_BY_CITY).fillna("")

    # drop empty name/address
    df = df[(df["name"] != "") & (df["address"] != "")]

    # chỉ giữ 2 loại
    df = df[df["category"].isin(["coworking space", "quán cà phê làm việc"])].copy()

    # dedupe lần cuối theo place/data id
    for col in ["data_id", "place_id", "data_cid"]:
        if col not in df.columns:
            df[col] = ""

    df["__k"] = (
        df["data_id"].fillna("").astype(str) + "|" +
        df["place_id"].fillna("").astype(str) + "|" +
        df["data_cid"].fillna("").astype(str) + "|" +
        df["name"].map(_norm) + "|" + df["address"].map(_norm)
    )
    df = df.drop_duplicates("__k").drop(columns=["__k"])

    # sắp xếp
    sort_cols = [c for c in ["region", "city", "category", "brand", "name"] if c in df.columns]
    df = df.sort_values(sort_cols, na_position="last")

    return df.reset_index(drop=True)


# ----------------------------
# Excel Export
# ----------------------------
DISPLAY_COLS = {
    "name": "Tên địa điểm",
    "category": "Loại hình",
    "brand": "Thương hiệu/Chuỗi",
    "address": "Địa chỉ",
    "city": "Tỉnh/Thành phố",
    "region": "Khu vực",
    "google_maps_link": "Link Google Maps",
    "website_or_social": "Website/MXH",
    "latitude": "Vĩ độ",
    "longitude": "Kinh độ",
    "notes": "Ghi chú",
    "image_url": "Ảnh (URL)",
    # raw
    "phone": "SĐT",
    "rating": "Điểm rating",
    "reviews_count": "Số review",
    "source_query": "Query nguồn",
    "place_id": "Place ID",
    "data_id": "Data ID",
    "data_cid": "CID",
}


def export_excel_bytes(
    df: pd.DataFrame,
    mode: str = "1 file (2 sheet: HN + HCM)",
    include_raw_cols: bool = False,
    filename: str = "coworking_cafe.xlsx",
) -> bytes:
    df = postprocess_dataframe(df)

    base_cols = [
        "name", "category", "brand", "address", "city", "region",
        "google_maps_link", "website_or_social","image_url",
        "latitude", "longitude", "notes",
    ]
    raw_cols = ["phone", "rating", "reviews_count", "source_query", "place_id", "data_id", "data_cid"]
    cols = base_cols + (raw_cols if include_raw_cols else [])

    # keep only existing columns
    df_out = df[[c for c in cols if c in df.columns]].copy()

    # rename to Vietnamese display headers
    df_out = df_out.rename(columns={k: v for k, v in DISPLAY_COLS.items() if k in df_out.columns})

    # convenient names after rename
    COL_CITY = DISPLAY_COLS.get("city", "Tỉnh/Thành phố")
    COL_REGION = DISPLAY_COLS.get("region", "Khu vực")

    bio = BytesIO()
    with pd.ExcelWriter(bio, engine="openpyxl") as writer:
        # Mode: theo Tỉnh/TP
        if mode.startswith("Theo tỉnh/TP"):
            if COL_CITY not in df_out.columns:
                df_out.to_excel(writer, index=False, sheet_name="data")
                _style_sheet(writer.sheets["data"], df_out.columns.tolist())
            else:
                for city in df_out[COL_CITY].dropna().unique().tolist():
                    sub = df_out[df_out[COL_CITY] == city].copy()
                    if sub.empty:
                        continue
                    sheet = _safe_sheet_name(city)
                    sub.to_excel(writer, index=False, sheet_name=sheet)
                    _style_sheet(writer.sheets[sheet], sub.columns.tolist())

        # Mode: theo Khu vực
        elif mode.startswith("Theo khu vực"):
            if COL_REGION not in df_out.columns:
                df_out.to_excel(writer, index=False, sheet_name="data")
                _style_sheet(writer.sheets["data"], df_out.columns.tolist())
            else:
                for region in df_out[COL_REGION].dropna().unique().tolist():
                    sub = df_out[df_out[COL_REGION] == region].copy()
                    if sub.empty:
                        continue
                    sheet = _safe_sheet_name(region)
                    sub.to_excel(writer, index=False, sheet_name=sheet)
                    _style_sheet(writer.sheets[sheet], sub.columns.tolist())

        # Giữ lại mode cũ của bạn
        elif mode.startswith("1 file (2 sheet"):
            for city in ["TP.HCM", "Hà Nội", "Hải Phòng", "Quảng Ninh"]:
                if COL_CITY not in df_out.columns:
                    continue
                sub = df_out[df_out[COL_CITY] == city].copy()
                if sub.empty:
                    continue
                sheet = _safe_sheet_name(city)
                sub.to_excel(writer, index=False, sheet_name=sheet)
                _style_sheet(writer.sheets[sheet], sub.columns.tolist())

        else:
            df_out.to_excel(writer, index=False, sheet_name="data")
            _style_sheet(writer.sheets["data"], df_out.columns.tolist())

    return bio.getvalue()


def _style_sheet(ws, columns: List[str]) -> None:
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2F5597")  # xanh đậm
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    wrap_cols = {"Địa chỉ", "Ghi chú"}
    url_cols = {"Link Google Maps", "Website/MXH"}

    # set widths
    for col_idx, col_name in enumerate(columns, start=1):
        letter = get_column_letter(col_idx)
        max_len = len(str(col_name))
        for r in range(2, min(ws.max_row, 250) + 1):
            v = ws[f"{letter}{r}"].value
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        width = min(max(12, max_len + 2), 55)
        ws.column_dimensions[letter].width = width

    # wrap + top align
    for col_name in columns:
        if col_name in wrap_cols:
            col_idx = columns.index(col_name) + 1
            letter = get_column_letter(col_idx)
            for r in range(2, ws.max_row + 1):
                ws[f"{letter}{r}"].alignment = Alignment(wrap_text=True, vertical="top")

    # hyperlink style
    link_font = Font(color="0563C1", underline="single")
    for col_name in columns:
        if col_name in url_cols:
            col_idx = columns.index(col_name) + 1
            letter = get_column_letter(col_idx)
            for r in range(2, ws.max_row + 1):
                cell = ws[f"{letter}{r}"]
                if isinstance(cell.value, str) and cell.value.startswith("http"):
                    cell.hyperlink = cell.value
                    cell.font = link_font


# CLI mode (optional)
if __name__ == "__main__":
    cities = ["TP.HCM", "Hà Nội", "Hải Phòng", "Quảng Ninh"]
    queries = build_queries(cities, use_core=True, use_brands=True, use_districts=True)
    rows = crawl_cities(cities, queries, max_pages_per_query=10, sleep_sec=0.3)
    df = postprocess_dataframe(pd.DataFrame(rows))

    # Examples:
    # mode = "Theo tỉnh/TP (mỗi sheet 1 tỉnh/TP)"
    # mode = "Theo khu vực (mỗi sheet 1 khu vực)"
    mode = "Theo tỉnh/TP (mỗi sheet 1 tỉnh/TP)"

    xlsx = export_excel_bytes(df, mode=mode, include_raw_cols=False, filename="coworking_cafe.xlsx")
    with open("coworking_cafe.xlsx", "wb") as f:
        f.write(xlsx)
    print("Done:", len(df), "rows -> coworking_cafe.xlsx")
